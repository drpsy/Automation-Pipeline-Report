"""
đọc dữ liệu từ file Excel và lưu trữ vào cơ sở dữ liệu PostgreSQL sau khi qua tiền xử lý
Qua đó biết được thông tin chi tiết về các mặt hàng còn tồn tại trong kho
"""


import pandas as pd
import os 
from sqlalchemy import create_engine
from openpyxl import load_workbook
import string
import re
from pyvi import ViUtils

# os.list(dir)



#read data from xlsx to dataframe
def iter_row(ws):
    for row in ws.iter_rows():
        yield[cell.value for cell in row[:]]

def get_schema(fname):
    wb = load_workbook(fname)
    schema = pd.DataFrame(iter_row(ws))
    schema.columns = schema.iloc[0]
    schema = schema.reindex(schema.index.drop(0))
    return schema


df = get_schema(".\data\inventory\inventory_24.1.xlsx")

engine = create_engine("engine = create_engine("postgresql://postgres:report_server@128.199.194.183:5432/report")")
   
pattern = "|".join([i for i in string.punctuation])

def remove_punct(test_str):
    for i in string.punctuation:
        if i in test_str: 
            test_str = test_str.replace(i, " ")
    return test_str

 
def preprocess(a):
    a = " ".join(remove_punct(a).split())
    a = " ".join(a.lower().split())
    a = ViUtils.remove_accents(a)
    a = a.decode()
    a = " ".join(a.lower().split())
    a = a.replace(" ", "_")
    return a


renames = dict()
for col in list(df.columns):
    renames[col] = preprocess(col)



def get_product_type(a):   
    a = a.split()
    return a[0]

def get_process_type(a):
    a = a.split()
    return a[1]
def get_size(a):
    a = a.split()
    try:
        return a[2]
    except:
        print(a)
        pass



df.to_sql("inventoryDetails", engine, if_exists = "replace")
